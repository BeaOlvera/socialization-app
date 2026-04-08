"""
Build check-in schedule Excel for the 12-month socialization journey.
Maps every check-in to phase, timing, who does it, and type.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Check-in Schedule"

# Styles
HEADER_FONT = Font(name="Segoe UI", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
SELF_FILL = PatternFill(start_color="EEEEF5", end_color="EEEEF5", fill_type="solid")
MGR_FILL = PatternFill(start_color="EAF4EF", end_color="EAF4EF", fill_type="solid")
HR_FILL = PatternFill(start_color="FBEAEC", end_color="FBEAEC", fill_type="solid")
BUDDY_FILL = PatternFill(start_color="FEF3E2", end_color="FEF3E2", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(
    left=Side(style="thin", color="E2E0DA"), right=Side(style="thin", color="E2E0DA"),
    top=Side(style="thin", color="E2E0DA"), bottom=Side(style="thin", color="E2E0DA"),
)

# Headers
headers = [
    ("Phase", 16), ("Month", 8), ("Week", 8), ("Day(s)", 12),
    ("Check-in Type", 22), ("Who Initiates", 18), ("Who Participates", 22),
    ("Format", 14), ("Duration", 10), ("Dimensions Covered", 16),
    ("Key Focus / Questions", 50), ("Output / Deliverable", 35),
]
for col, (h, w) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN
    ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = w

# Check-in data
# Format: (phase, month, week, days, type, initiator, participants, format, duration, dimensions, focus, output)
checkins = [
    # PRE-ARRIVAL
    ("Pre-arrival", "-", "W-2", "10 days before", "Welcome call", "HR Admin", "Newcomer + HR", "Video call", "20 min", "FIT, TIE", "Introduction, logistics, expectations, answer questions", "Newcomer feels expected; logistics confirmed"),
    ("Pre-arrival", "-", "W-1", "5 days before", "Manager intro call", "Manager", "Newcomer + Manager", "Video call", "30 min", "FIT, TIE", "Role overview, first week plan, team intro, manager's expectations", "Newcomer knows what to expect day 1; rapport started"),

    # ARRIVAL (Days 1-30)
    ("Arrival", "1", "W1", "Day 1", "Day-1 welcome check-in", "Manager", "Newcomer + Manager + Buddy", "In person", "30 min", "FIT, ACE, TIE", "Welcome, workspace setup, introductions, first-day plan", "Newcomer oriented; buddy relationship started"),
    ("Arrival", "1", "W1", "Day 3", "Buddy informal check-in", "Buddy", "Newcomer + Buddy", "Coffee / informal", "15 min", "TIE", "How are you settling in? Questions? Social integration", "Early concerns surfaced; social comfort"),
    ("Arrival", "1", "W1", "Day 5", "End-of-week 1 check-in", "Manager", "Newcomer + Manager", "1:1 meeting", "30 min", "FIT, ACE", "First impressions, clarity on role, tools working?, blockers", "Week 1 blockers resolved; expectations aligned"),
    ("Arrival", "1", "W2", "Day 10", "Buddy check-in", "Buddy", "Newcomer + Buddy", "Coffee / informal", "15 min", "TIE, ACE", "Social connections, team dynamics, any gaps in tools/access", "Social network expanding; practical gaps closed"),
    ("Arrival", "1", "W2", "Day 14", "2-week self check-in", "Newcomer (self)", "Newcomer", "App survey (Likert)", "10 min", "FIT, ACE, TIE", "5 questions per dimension: role clarity, tool mastery, belonging", "Self-assessment baseline score recorded"),
    ("Arrival", "1", "W3", "Day 18", "Manager 1:1", "Manager", "Newcomer + Manager", "1:1 meeting", "30 min", "FIT, ACE", "KPI discussion, training progress, feedback on first tasks", "KPIs understood; training gaps identified"),
    ("Arrival", "1", "W4", "Day 25", "Buddy check-in", "Buddy", "Newcomer + Buddy", "Informal", "15 min", "TIE", "Deeper social integration, cross-team connections", "Newcomer has connections beyond immediate team"),
    ("Arrival", "1", "W4", "Day 30", "Month 1 formal check-in (Self)", "Newcomer (self)", "Newcomer", "App survey (Likert) + AI interview", "15 min", "FIT, ACE, TIE", "Full Likert assessment + qualitative AI-guided reflection", "Month 1 self-score; qualitative insights captured"),
    ("Arrival", "1", "W4", "Day 30", "Month 1 formal check-in (Manager)", "Manager", "Manager", "App survey (Likert)", "10 min", "FIT, ACE, TIE", "Manager rates newcomer on same 15 questions; divergence flagged", "Manager score; self vs manager comparison available"),

    # INTEGRATION (Days 31-90)
    ("Integration", "2", "W5", "Day 35", "Buddy check-in", "Buddy", "Newcomer + Buddy", "Informal", "15 min", "TIE", "Social rituals, ERG/community involvement, any isolation signs", "Social integration on track"),
    ("Integration", "2", "W6", "Day 42", "Manager 1:1", "Manager", "Newcomer + Manager", "1:1 meeting", "30 min", "FIT, ACE", "First project review, RACI clarity, skill gap plan", "First deliverable feedback; development plan started"),
    ("Integration", "2", "W8", "Day 56", "Month 2 self check-in", "Newcomer (self)", "Newcomer", "App survey (Likert)", "10 min", "FIT, ACE, TIE", "Same 15 questions — tracking trend vs month 1", "Month 2 self-score; trend visible"),
    ("Integration", "3", "W9", "Day 63", "Buddy check-in", "Buddy", "Newcomer + Buddy", "Informal", "15 min", "TIE, ACE", "Cross-team collaboration, knowledge sharing, confidence", "Social + competence check"),
    ("Integration", "3", "W10", "Day 70", "Manager 1:1", "Manager", "Newcomer + Manager", "1:1 meeting", "30 min", "FIT, ACE", "90-day plan progress, training completion, performance early signs", "On track for 90-day milestone"),
    ("Integration", "3", "W12", "Day 84", "Month 3 formal check-in (Self)", "Newcomer (self)", "Newcomer", "App survey (Likert) + AI interview", "15 min", "FIT, ACE, TIE", "Full assessment + qualitative reflection on integration phase", "Month 3 self-score; qualitative insights"),
    ("Integration", "3", "W12", "Day 84", "Month 3 formal check-in (Manager)", "Manager", "Manager", "App survey (Likert)", "10 min", "FIT, ACE, TIE", "Manager assessment; divergence analysis vs self-scores", "Manager score; end-of-integration assessment"),
    ("Integration", "3", "W12", "Day 90", "HR 90-day review", "HR Admin", "Newcomer + Manager + HR", "Meeting", "45 min", "FIT, ACE, TIE", "Formal 90-day review: progress, blockers, development plan, risk flags", "90-day report; action plan if at-risk"),

    # ADJUSTMENT (Months 4-6)
    ("Adjustment", "4", "W14", "Day 98", "Manager 1:1", "Manager", "Newcomer + Manager", "1:1 meeting", "30 min", "FIT, ACE", "Quarterly priorities, independence level, process improvements", "Q2 goals set; growing autonomy confirmed"),
    ("Adjustment", "4", "W16", "Day 112", "Month 4 self check-in", "Newcomer (self)", "Newcomer", "App survey (Likert)", "10 min", "FIT, ACE, TIE", "Tracking trend; focus on autonomy and belonging", "Month 4 self-score"),
    ("Adjustment", "5", "W18", "Day 126", "Buddy check-in", "Buddy", "Newcomer + Buddy", "Informal", "15 min", "TIE", "Network depth, informal leadership, values alignment", "Social embedding progressing"),
    ("Adjustment", "5", "W20", "Day 140", "Month 5 self check-in", "Newcomer (self)", "Newcomer", "App survey (Likert)", "10 min", "FIT, ACE, TIE", "Midpoint assessment — halfway through first year", "Month 5 self-score"),
    ("Adjustment", "6", "W22", "Day 154", "Manager 1:1", "Manager", "Newcomer + Manager", "1:1 meeting", "30 min", "FIT, ACE", "Mid-year performance preview, results review, development", "Mid-year readiness confirmed"),
    ("Adjustment", "6", "W24", "Day 168", "Month 6 formal check-in (Self)", "Newcomer (self)", "Newcomer", "App survey (Likert) + AI interview", "15 min", "FIT, ACE, TIE", "Full assessment + reflection on adjustment phase", "Month 6 self-score; qualitative insights"),
    ("Adjustment", "6", "W24", "Day 168", "Month 6 formal check-in (Manager)", "Manager", "Manager", "App survey (Likert)", "10 min", "FIT, ACE, TIE", "Manager mid-year assessment; divergence check", "Manager score; mid-year comparison"),
    ("Adjustment", "6", "W24", "Day 180", "HR 6-month review", "HR Admin", "Newcomer + Manager + HR", "Meeting", "45 min", "FIT, ACE, TIE", "Formal 6-month review: trajectory, retention risk, development plan update", "6-month report; updated action plan"),

    # STABILIZATION (Months 7-9)
    ("Stabilization", "7", "W28", "Day 196", "Manager 1:1", "Manager", "Newcomer + Manager", "1:1 meeting", "30 min", "FIT, ACE", "Domain expertise, cross-functional leadership, mentoring readiness", "Stabilization goals confirmed"),
    ("Stabilization", "7", "W30", "Day 210", "Month 7 self check-in", "Newcomer (self)", "Newcomer", "App survey (Likert)", "10 min", "FIT, ACE, TIE", "Tracking trend; focus on expertise and trust", "Month 7 self-score"),
    ("Stabilization", "8", "W34", "Day 238", "Month 8 self check-in", "Newcomer (self)", "Newcomer", "App survey (Likert)", "10 min", "FIT, ACE, TIE", "Consistency check — are scores stabilizing?", "Month 8 self-score"),
    ("Stabilization", "9", "W36", "Day 252", "Manager 1:1", "Manager", "Newcomer + Manager", "1:1 meeting", "30 min", "FIT, ACE", "Performance appraisal prep, development path, next-year goals preview", "Appraisal readiness; development plan"),
    ("Stabilization", "9", "W38", "Day 266", "Month 9 formal check-in (Self)", "Newcomer (self)", "Newcomer", "App survey (Likert) + AI interview", "15 min", "FIT, ACE, TIE", "Full assessment + reflection on stabilization", "Month 9 self-score; qualitative insights"),
    ("Stabilization", "9", "W38", "Day 266", "Month 9 formal check-in (Manager)", "Manager", "Manager", "App survey (Likert)", "10 min", "FIT, ACE, TIE", "Manager assessment; near-final divergence check", "Manager score; stabilization comparison"),

    # EMBEDDING (Months 10-12)
    ("Embedding", "10", "W40", "Day 280", "Manager 1:1", "Manager", "Newcomer + Manager", "1:1 meeting", "30 min", "FIT, ACE", "Strategic contribution, next-year goals, career path", "Year-end planning started"),
    ("Embedding", "10", "W42", "Day 294", "Month 10 self check-in", "Newcomer (self)", "Newcomer", "App survey (Likert)", "10 min", "FIT, ACE, TIE", "Near-final self-assessment; identity within org", "Month 10 self-score"),
    ("Embedding", "11", "W46", "Day 322", "Month 11 self check-in", "Newcomer (self)", "Newcomer", "App survey (Likert)", "10 min", "FIT, ACE, TIE", "Penultimate check — confidence and belonging", "Month 11 self-score"),
    ("Embedding", "12", "W48", "Day 336", "Manager 1:1", "Manager", "Newcomer + Manager", "1:1 meeting", "30 min", "FIT, ACE, TIE", "Final performance review, 12-month reflection, transition to regular employee", "Year-end appraisal input ready"),
    ("Embedding", "12", "W50", "Day 350", "Month 12 formal check-in (Self)", "Newcomer (self)", "Newcomer", "App survey (Likert) + AI interview", "15 min", "FIT, ACE, TIE", "Final full assessment + qualitative reflection on entire journey", "Month 12 final self-score; qualitative summary"),
    ("Embedding", "12", "W50", "Day 350", "Month 12 formal check-in (Manager)", "Manager", "Manager", "App survey (Likert)", "10 min", "FIT, ACE, TIE", "Manager final assessment; full year divergence analysis", "Manager final score; 12-month comparison"),
    ("Embedding", "12", "W52", "Day 365", "HR 12-month review", "HR Admin", "Newcomer + Manager + HR", "Meeting", "60 min", "FIT, ACE, TIE", "Final review: full trajectory, socialization outcome, transition to 'regular employee', retention plan", "12-month report; socialization complete"),
    ("Embedding", "12", "W52", "Day 365", "Socialization completion celebration", "HR Admin", "Newcomer + Team", "Team event", "30 min", "TIE", "Recognition, certificate, team celebration of successful onboarding", "Certificate issued; socialization journey closes"),
]

TYPE_FILLS = {
    "self": SELF_FILL,
    "manager": MGR_FILL,
    "hr": HR_FILL,
    "buddy": BUDDY_FILL,
}

for i, row in enumerate(checkins, start=2):
    for col, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.alignment = WRAP
        cell.border = THIN

    # Color-code by who initiates
    initiator = row[5].lower()
    fill = None
    if "newcomer" in initiator or "self" in initiator:
        fill = SELF_FILL
    elif "manager" in initiator:
        fill = MGR_FILL
    elif "hr" in initiator:
        fill = HR_FILL
    elif "buddy" in initiator:
        fill = BUDDY_FILL

    if fill:
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).fill = fill

# Summary sheet
ws2 = wb.create_sheet("Summary")
ws2.cell(row=1, column=1, value="Check-in Type Summary").font = Font(bold=True, size=12)

summary = [
    ("", "Count", "Frequency"),
    ("Self check-in (Likert only)", "8", "Months 2, 4, 5, 7, 8, 10, 11 + W2"),
    ("Self check-in (Likert + AI interview)", "5", "Months 1, 3, 6, 9, 12"),
    ("Manager check-in (Likert)", "5", "Months 1, 3, 6, 9, 12"),
    ("Manager 1:1 meetings", "9", "W1, W3, W6, W10, W14, W22, W28, W36, W40, W48"),
    ("Buddy informal check-ins", "6", "Days 3, 10, 25, 35, 63, 126"),
    ("HR formal reviews", "3", "Day 90, Day 180, Day 365"),
    ("Other (welcome call, celebration)", "3", "Pre-arrival + Day 1 + Day 365"),
    ("", "", ""),
    ("TOTAL CHECK-INS", str(len(checkins)), "39 touchpoints over 12 months"),
]

for i, (a, b, c) in enumerate(summary, start=3):
    ws2.cell(row=i, column=1, value=a).font = Font(bold=(a == "TOTAL CHECK-INS"))
    ws2.cell(row=i, column=2, value=b)
    ws2.cell(row=i, column=3, value=c)

ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 50

output = "Socialization_Checkin_Schedule.xlsx"
wb.save(output)
print(f"Done! {len(checkins)} check-ins written to {output}")
