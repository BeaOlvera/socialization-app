"""
Build the complete 12-month socialization activity calendar.
Outputs an Excel file with ~200 activities mapped to days, dimensions, subdimensions.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Socialization Calendar"

# ── Styles ──
HEADER_FONT = Font(name="Segoe UI", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
FIT_FILL = PatternFill(start_color="EEEEF5", end_color="EEEEF5", fill_type="solid")
ACE_FILL = PatternFill(start_color="EAF4EF", end_color="EAF4EF", fill_type="solid")
TIE_FILL = PatternFill(start_color="FBEAEC", end_color="FBEAEC", fill_type="solid")
PRE_FILL = PatternFill(start_color="FEF3E2", end_color="FEF3E2", fill_type="solid")
DIM_FILLS = {"FIT": FIT_FILL, "ACE": ACE_FILL, "TIE": TIE_FILL}
PHASE_FILLS = {
    "Pre-arrival": PRE_FILL,
    "Arrival": PatternFill(start_color="F5F4F0", end_color="F5F4F0", fill_type="solid"),
    "Integration": PatternFill(start_color="F5F4F0", end_color="F5F4F0", fill_type="solid"),
    "Adjustment": PatternFill(start_color="F5F4F0", end_color="F5F4F0", fill_type="solid"),
    "Stabilization": PatternFill(start_color="F5F4F0", end_color="F5F4F0", fill_type="solid"),
    "Embedding": PatternFill(start_color="F5F4F0", end_color="F5F4F0", fill_type="solid"),
}
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(
    left=Side(style="thin", color="E2E0DA"), right=Side(style="thin", color="E2E0DA"),
    top=Side(style="thin", color="E2E0DA"), bottom=Side(style="thin", color="E2E0DA"),
)

# ── Headers ──
headers = [
    ("Phase", 14), ("Week", 6), ("Day(s)", 10), ("Dimension", 8), ("Subdimension", 20),
    ("Activity", 50), ("Who", 18), ("Est. Time", 10), ("Builds On", 30), ("Output / Deliverable", 30),
]
for col, (h, w) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN
    ws.column_dimensions[get_column_letter(col)].width = w

# ── Activity data ──
# Format: (phase, week, days, dim, subdim, activity, who, time, builds_on, output)

activities = [
    # ═══════════════════════════════════════════════════════════
    # PRE-ARRIVAL (Week -2 to Day 0)
    # ═══════════════════════════════════════════════════════════
    ("Pre-arrival", "-2", "10 days before", "FIT", "Job Description",
     "Read the full job description. Highlight 3 things you're excited about and 3 that feel unclear.",
     "Newcomer alone", "30 min", "—", "Annotated job description with questions list"),

    ("Pre-arrival", "-2", "10 days before", "FIT", "Org Chart",
     "Review the org chart document sent by HR. Identify your manager, skip-level, and 3 key departments you'll interact with.",
     "Newcomer alone", "20 min", "—", "Mental map of key people and structure"),

    ("Pre-arrival", "-2", "10 days before", "ACE", "Training Plan",
     "Read the welcome pack and 30-60-90 day plan. Note which tools and systems are mentioned.",
     "Newcomer alone", "30 min", "—", "List of tools to expect; questions about milestones"),

    ("Pre-arrival", "-1", "5 days before", "TIE", "Company Values",
     "Read the culture guide. For each company value, think of a situation from your past experience where you demonstrated it.",
     "Newcomer alone", "30 min", "—", "Personal examples linked to each value"),

    ("Pre-arrival", "-1", "5 days before", "TIE", "Buddy / Mentor",
     "Your buddy will send you a short intro message. Reply with a brief intro about yourself and one thing you're curious about.",
     "Newcomer + Buddy", "15 min", "—", "Ice-breaking message exchange"),

    ("Pre-arrival", "-1", "3 days before", "FIT", "Job Description",
     "Write down your top 5 questions about the role to ask your manager in the first 1:1.",
     "Newcomer alone", "20 min", "Annotated job description", "5 prepared questions for manager"),

    ("Pre-arrival", "-1", "3 days before", "ACE", "Tools & Systems",
     "Set up any accounts sent to you (email, Slack, VPN). Test that they work.",
     "Newcomer alone", "30 min", "—", "All pre-arrival accounts active"),

    ("Pre-arrival", "-1", "1 day before", "FIT", "Strategic Plan",
     "Watch the company strategy video (if provided). Note how your role might connect to the bigger picture.",
     "Newcomer alone", "20 min", "Org chart review", "Initial sense of strategic context"),

    # ═══════════════════════════════════════════════════════════
    # ARRIVAL — Week 1 (Days 1-5)
    # ═══════════════════════════════════════════════════════════
    ("Arrival", "1", "Day 1", "TIE", "Team Rituals",
     "Attend welcome breakfast with your team. Introduce yourself: name, background, one fun fact.",
     "Newcomer + Team", "1 hour", "—", "First face-to-face with team"),

    ("Arrival", "1", "Day 1", "ACE", "Tools & Systems",
     "IT setup session: laptop, badge, tool accounts. Verify access to email, Slack, calendar, project tools.",
     "Newcomer + IT", "2 hours", "Pre-arrival account setup", "Full system access confirmed"),

    ("Arrival", "1", "Day 1", "TIE", "Buddy / Mentor",
     "Meet your buddy for coffee. Ask them: What surprised you most when you joined? What's the unwritten rule #1?",
     "Newcomer + Buddy", "30 min", "Pre-arrival intro message", "Buddy relationship started; 2 insider tips noted"),

    ("Arrival", "1", "Day 1", "FIT", "Job Description",
     "1:1 with manager: walk through the job description together. Ask your 5 prepared questions. Clarify top 3 priorities.",
     "Newcomer + Manager", "45 min", "5 prepared questions", "Shared understanding of role priorities"),

    ("Arrival", "1", "Day 1", "TIE", "Company Values",
     "Attend company All-Hands (if scheduled). Observe which values are mentioned and how people interact.",
     "Newcomer + Company", "30 min", "Culture guide reading", "First-hand observation of values in action"),

    ("Arrival", "1", "Day 2", "FIT", "Org Chart",
     "HR orientation: walk through the org chart live. Ask: Who are the 3 people outside my team I'll work with most?",
     "Newcomer + HR", "30 min", "Pre-arrival org chart review", "Annotated org chart with key contacts marked"),

    ("Arrival", "1", "Day 2", "ACE", "SOPs & Playbooks",
     "Tour the internal wiki/Notion/SharePoint. Bookmark the 5 most relevant pages for your role.",
     "Newcomer + Buddy", "30 min", "Tool access confirmed", "5 key resources bookmarked"),

    ("Arrival", "1", "Day 2", "TIE", "Buddy / Mentor",
     "Buddy introduces you to 2-3 team members one-on-one. 5 min each — learn what they do and how you'll interact.",
     "Newcomer + Buddy + Peers", "20 min", "Buddy coffee meeting", "3 new faces known by name + role"),

    ("Arrival", "1", "Day 3", "ACE", "Training Plan",
     "Review your 30-60-90 day plan with manager. Confirm: What does 'success' look like at Day 30?",
     "Newcomer + Manager", "30 min", "Role clarification (Day 1)", "Agreed Day 30 milestones written down"),

    ("Arrival", "1", "Day 3", "FIT", "KPIs / OKRs",
     "Manager explains your KPIs and how they're measured. Ask: Which KPI matters most? How will I know I'm on track?",
     "Newcomer + Manager", "30 min", "Role priorities clarified", "KPIs documented with measurement criteria"),

    ("Arrival", "1", "Day 3", "ACE", "Tools & Systems",
     "Deep-dive into your #1 daily tool (CRM, project tool, etc.). Complete a tutorial or shadow a colleague using it.",
     "Newcomer + Peer", "1 hour", "System access confirmed", "Can navigate primary tool independently"),

    ("Arrival", "1", "Day 4", "FIT", "RACI Matrix",
     "Ask your manager: For my main deliverables, who decides, who executes, who needs to be consulted? Draft a mini-RACI.",
     "Newcomer + Manager", "30 min", "KPIs documented", "Draft RACI for top 3 processes"),

    ("Arrival", "1", "Day 4", "TIE", "Employee Communities",
     "Browse Slack channels, ERGs, interest groups. Join 2 that interest you. Say hello in one.",
     "Newcomer alone", "20 min", "Slack access", "Joined 2 communities; first message sent"),

    ("Arrival", "1", "Day 4", "ACE", "Performance Appraisal",
     "HR explains the performance review process: timeline, format, competencies rated. Note when your first review is.",
     "Newcomer + HR", "20 min", "—", "Review calendar marked; competencies list saved"),

    ("Arrival", "1", "Day 5", "TIE", "Team Rituals",
     "Attend Friday team lunch (or equivalent ritual). Don't skip it — this is where informal bonds form.",
     "Newcomer + Team", "1 hour", "—", "Participated in first team social ritual"),

    ("Arrival", "1", "Day 5", "FIT", "Strategic Plan",
     "Read the latest strategy deck or annual priorities. Write one sentence: 'My role supports the strategy by...'",
     "Newcomer alone", "30 min", "Strategy video (pre-arrival)", "One-sentence role-strategy link"),

    ("Arrival", "1", "Day 5", "ACE", "Skills Matrix",
     "Self-assess: For each skill your role requires, rate yourself 1-5. Identify top 2 gaps to discuss with manager.",
     "Newcomer alone", "30 min", "KPIs and role clarity", "Self-assessment with 2 priority gaps identified"),

    ("Arrival", "1", "Day 5", "FIT", "Job Description",
     "End-of-week 1:1 with manager. Share: 'Here's what I understand my role to be' — get confirmation or correction.",
     "Newcomer + Manager", "30 min", "All Week 1 activities", "Validated role understanding; misunderstandings corrected"),

    # ═══════════════════════════════════════════════════════════
    # ARRIVAL — Week 2 (Days 6-10)
    # ═══════════════════════════════════════════════════════════
    ("Arrival", "2", "Day 6-7", "FIT", "Org Chart",
     "Set up intro meetings with 3 key cross-functional contacts identified in Week 1. 20 min each: learn their role, how you'll collaborate.",
     "Newcomer + Cross-functional", "1 hour total", "Annotated org chart", "3 cross-functional relationships initiated"),

    ("Arrival", "2", "Day 6-7", "ACE", "SOPs & Playbooks",
     "Execute one simple process end-to-end using the documented SOP. Note where the SOP is outdated or unclear.",
     "Newcomer alone", "1 hour", "Wiki bookmarks", "First process completed; SOP feedback notes"),

    ("Arrival", "2", "Day 8", "FIT", "RACI Matrix",
     "Test your RACI understanding: before starting a task, state who you think should approve it. Check with manager.",
     "Newcomer + Manager", "15 min", "Draft RACI", "RACI validated on a real task"),

    ("Arrival", "2", "Day 8-9", "ACE", "Training Plan",
     "Complete first mandatory training module (compliance, security, tool certification — whatever is assigned).",
     "Newcomer alone", "1-2 hours", "Training plan review", "First training module completed"),

    ("Arrival", "2", "Day 9", "TIE", "Pulse Surveys",
     "Complete your first pulse survey (Week 2 check-in). Be honest — this data helps HR support you.",
     "Newcomer alone", "10 min", "—", "First pulse survey submitted"),

    ("Arrival", "2", "Day 9-10", "TIE", "Company Values",
     "Ask a colleague: 'Which company value do you see lived most strongly here? Can you give me an example?' Note the gap between stated and lived values.",
     "Newcomer + Peer", "15 min", "Culture guide + All-Hands observation", "Real example of values in action collected"),

    ("Arrival", "2", "Day 10", "ACE", "Tools & Systems",
     "Can you complete your daily workflow without asking for help on tools? Self-test. Flag remaining access issues to IT.",
     "Newcomer alone", "30 min", "Tool deep-dives", "Tool independence self-assessment; IT tickets filed"),

    # ═══════════════════════════════════════════════════════════
    # ARRIVAL — Week 3-4 (Days 11-30)
    # ═══════════════════════════════════════════════════════════
    ("Arrival", "3", "Day 11-15", "FIT", "KPIs / OKRs",
     "Review your KPI dashboard/tracker for the first time. Understand where data comes from. Ask manager to walk through one metric.",
     "Newcomer + Manager", "30 min", "KPIs documented", "Can read own KPI dashboard independently"),

    ("Arrival", "3", "Day 11-15", "ACE", "Training Plan",
     "Shadow a colleague doing the work you'll own. Take notes on their process, shortcuts, and quality standards.",
     "Newcomer + Peer", "2 hours", "First SOP execution", "Shadowing notes with process insights"),

    ("Arrival", "3", "Day 11-15", "TIE", "Buddy / Mentor",
     "Week 3 buddy check-in. Ask: Am I missing any informal norms? Is there anyone I should meet that I haven't?",
     "Newcomer + Buddy", "20 min", "Week 1-2 buddy meetings", "Updated list of people to connect with"),

    ("Arrival", "3", "Day 11-15", "FIT", "Strategic Plan",
     "Attend a team meeting where strategy is discussed. Note how decisions reference company priorities.",
     "Newcomer + Team", "1 hour", "Strategy one-liner", "Observed strategic decision-making in action"),

    ("Arrival", "4", "Day 16-20", "ACE", "Skills Matrix",
     "Discuss your self-assessment with manager. Agree on top 2 skill gaps and a plan to close them (training, mentoring, practice).",
     "Newcomer + Manager", "30 min", "Self-assessment (Day 5)", "Agreed development plan for 2 priority skills"),

    ("Arrival", "4", "Day 16-20", "TIE", "Employee Communities",
     "Attend one community/ERG event or virtual meetup. Meet 2 people from outside your department.",
     "Newcomer + Community", "1 hour", "Joined 2 communities", "2 new cross-department contacts"),

    ("Arrival", "4", "Day 16-20", "FIT", "RACI Matrix",
     "Proactively identify one potential role overlap or boundary issue. Raise it with manager: 'I think X and Y might clash — who owns this?'",
     "Newcomer + Manager", "15 min", "RACI validation", "One boundary issue clarified proactively"),

    ("Arrival", "4", "Day 21-25", "ACE", "SOPs & Playbooks",
     "Complete a second process end-to-end independently. Compare your output to the quality standard. Self-assess.",
     "Newcomer alone", "1-2 hours", "First SOP execution + shadowing", "Second deliverable completed; quality self-check done"),

    ("Arrival", "4", "Day 21-25", "TIE", "Team Rituals",
     "By now you should be attending all recurring meetings. Check: Am I on every calendar invite I should be? Any I'm missing?",
     "Newcomer alone", "10 min", "Week 1-3 ritual attendance", "Calendar audit complete; missing invites requested"),

    ("Arrival", "4", "Day 26-30", "FIT", "Job Description",
     "End of Arrival reflection with manager: 'Here's what I own, here's what I don't, here's where I'm still unclear.' Get sign-off.",
     "Newcomer + Manager", "30 min", "All Arrival FIT activities", "Role clarity confirmed for Integration phase"),

    ("Arrival", "4", "Day 26-30", "ACE", "Performance Appraisal",
     "Ask manager: How am I doing so far? What should I do more of / less of before my first formal review?",
     "Newcomer + Manager", "20 min", "Review process explanation (Day 4)", "Informal early feedback received"),

    ("Arrival", "4", "Day 26-30", "TIE", "Pulse Surveys",
     "Complete Month 1 check-in: quantitative (Likert) + qualitative (AI interview). Reflect honestly on belonging.",
     "Newcomer alone", "15 min", "All Arrival TIE activities", "Month 1 check-in submitted"),

    # ═══════════════════════════════════════════════════════════
    # INTEGRATION — Week 5-8 (Days 31-60)
    # ═══════════════════════════════════════════════════════════
    ("Integration", "5", "Day 31-35", "FIT", "KPIs / OKRs",
     "Set your personal OKRs for the quarter with manager. Align them to team and company OKRs. Agree on 3 measurable goals.",
     "Newcomer + Manager", "45 min", "KPI dashboard familiarity", "Personal quarterly OKRs documented"),

    ("Integration", "5", "Day 31-35", "ACE", "Training Plan",
     "Take ownership of your first project or deliverable end-to-end. Define scope, timeline, stakeholders.",
     "Newcomer + Manager", "1 hour", "Shadowing + 2 SOPs done", "First project plan with timeline"),

    ("Integration", "5", "Day 31-35", "TIE", "Buddy / Mentor",
     "Transition buddy relationship from 'survival guide' to 'sounding board'. Discuss a real work challenge with them.",
     "Newcomer + Buddy", "30 min", "3 buddy meetings", "Buddy as advisory relationship established"),

    ("Integration", "6", "Day 36-42", "FIT", "Org Chart",
     "Map your stakeholder network: Who do I influence? Who influences me? Who do I need but don't have yet? Share with manager.",
     "Newcomer alone + Manager review", "45 min", "Cross-functional intros", "Stakeholder map document"),

    ("Integration", "6", "Day 36-42", "ACE", "Tools & Systems",
     "Learn an advanced feature of your primary tool that saves time (automations, templates, shortcuts). Share it with a peer.",
     "Newcomer alone", "30 min", "Tool independence achieved", "One productivity tip mastered and shared"),

    ("Integration", "6", "Day 36-42", "TIE", "Company Values",
     "Identify one situation this week where a company value guided (or should have guided) a decision. Discuss with buddy.",
     "Newcomer + Buddy", "15 min", "Values observation + examples", "Values-in-action applied to real work"),

    ("Integration", "7", "Day 43-49", "FIT", "RACI Matrix",
     "Lead a handoff or collaboration with another team. Clarify roles before starting: 'I'll do X, you do Y, Z approves.'",
     "Newcomer + Cross-functional", "30 min", "RACI validated", "First cross-functional handoff led cleanly"),

    ("Integration", "7", "Day 43-49", "ACE", "SOPs & Playbooks",
     "Identify one SOP that's outdated or missing. Draft an improvement or create a new mini-playbook. Share with team.",
     "Newcomer alone", "1 hour", "3 SOPs executed", "One SOP improved or created"),

    ("Integration", "7", "Day 43-49", "TIE", "Team Rituals",
     "Volunteer to present something in a team meeting (a finding, an idea, a quick update). Step into visibility.",
     "Newcomer + Team", "15 min prep + meeting", "Regular meeting attendance", "First team presentation delivered"),

    ("Integration", "8", "Day 50-56", "ACE", "Performance Appraisal",
     "Prepare for first formal performance touchpoint: self-assess against competencies. Bring evidence of impact.",
     "Newcomer alone", "30 min", "Informal feedback (Day 26-30)", "Self-assessment with evidence prepared"),

    ("Integration", "8", "Day 50-56", "FIT", "Strategic Plan",
     "Present to manager: 'Here's how my current projects connect to our strategic priorities.' Get alignment.",
     "Newcomer + Manager", "20 min", "Strategy observation + OKRs", "Work-to-strategy link validated by manager"),

    ("Integration", "8", "Day 50-56", "TIE", "Employee Communities",
     "Organise or attend a cross-department social activity (coffee roulette, lunch group, after-work event).",
     "Newcomer + Cross-department", "1 hour", "Community membership", "Network extended beyond own team"),

    ("Integration", "8", "Day 57-60", "TIE", "Pulse Surveys",
     "Complete Month 2 check-in: quantitative + qualitative. Reflect on growth since Month 1.",
     "Newcomer alone", "15 min", "Month 1 check-in", "Month 2 check-in submitted"),

    # ═══════════════════════════════════════════════════════════
    # INTEGRATION — Week 9-13 (Days 61-90)
    # ═══════════════════════════════════════════════════════════
    ("Integration", "9-10", "Day 61-70", "ACE", "Training Plan",
     "Deliver your first project results. Present outcomes to manager and stakeholders. Collect feedback.",
     "Newcomer + Stakeholders", "2 hours", "First project plan", "First project delivered with measurable outcomes"),

    ("Integration", "9-10", "Day 61-70", "FIT", "KPIs / OKRs",
     "First formal KPI review with manager. Compare results to targets. Discuss what's working and what to adjust.",
     "Newcomer + Manager", "30 min", "Personal OKRs + KPI dashboard", "First KPI review completed; adjustments agreed"),

    ("Integration", "9-10", "Day 61-70", "ACE", "Skills Matrix",
     "Progress check on skill development plan. Have you closed one gap? Identify next gap to work on.",
     "Newcomer + Manager", "20 min", "Agreed dev plan (Week 4)", "Skill gap #1 closed or progressing; gap #2 identified"),

    ("Integration", "11-12", "Day 71-84", "TIE", "Buddy / Mentor",
     "Buddy mid-point review: What's going well? What do I wish I'd known earlier? Is there anyone left to meet?",
     "Newcomer + Buddy", "20 min", "Ongoing buddy check-ins", "Buddy relationship health check"),

    ("Integration", "11-12", "Day 71-84", "FIT", "RACI Matrix",
     "Review and update your RACI document. Share it with your team. Are there still grey areas?",
     "Newcomer + Team", "30 min", "Cross-functional handoff experience", "RACI v2 shared with team"),

    ("Integration", "13", "Day 85-90", "TIE", "Pulse Surveys",
     "Complete Month 3 check-in. This is the Integration milestone — reflect on how far you've come.",
     "Newcomer alone", "15 min", "Month 2 check-in", "Month 3 check-in submitted; Integration phase closed"),

    ("Integration", "13", "Day 85-90", "FIT", "Job Description",
     "90-day role review with manager: Has the role evolved from the original description? Update if needed.",
     "Newcomer + Manager", "30 min", "All Integration FIT activities", "Job description updated or confirmed for Adjustment phase"),

    ("Integration", "13", "Day 85-90", "ACE", "Performance Appraisal",
     "First formal performance review (if scheduled at 90 days). Receive structured feedback on competencies.",
     "Newcomer + Manager + HR", "1 hour", "Self-assessment + evidence", "90-day performance review completed"),

    # ═══════════════════════════════════════════════════════════
    # ADJUSTMENT — Month 4-6 (Days 91-180)
    # ═══════════════════════════════════════════════════════════
    ("Adjustment", "14-17", "Month 4", "FIT", "Strategic Plan",
     "Contribute to quarterly planning: propose one initiative aligned with strategic priorities. Present to team.",
     "Newcomer + Team", "2 hours", "Work-strategy alignment validated", "First strategic initiative proposed"),

    ("Adjustment", "14-17", "Month 4", "ACE", "Training Plan",
     "Own a major project independently from start to finish. No hand-holding — manager available but not leading.",
     "Newcomer (with Manager oversight)", "Ongoing", "First project delivered", "Full project ownership demonstrated"),

    ("Adjustment", "14-17", "Month 4", "TIE", "Team Rituals",
     "Take on a team ritual role (facilitate a meeting, organise the next team lunch, lead a retrospective).",
     "Newcomer + Team", "1 hour", "Regular ritual participation", "Led or organised one team ritual"),

    ("Adjustment", "14-17", "Month 4", "TIE", "Pulse Surveys",
     "Complete Month 4 check-in. Focus on: Do I feel I belong? Do people seek me out?",
     "Newcomer alone", "15 min", "Month 3 check-in", "Month 4 check-in submitted"),

    ("Adjustment", "18-22", "Month 5", "FIT", "KPIs / OKRs",
     "Mid-year OKR review. Are you hitting targets? Recalibrate if needed. Propose stretch goals.",
     "Newcomer + Manager", "30 min", "Q1 KPI review", "Mid-year OKR review completed; stretch goals set"),

    ("Adjustment", "18-22", "Month 5", "ACE", "SOPs & Playbooks",
     "Become the go-to person for at least one process. Others come to you for questions about it.",
     "Newcomer + Team", "Ongoing", "SOP contributions", "Recognised as process expert in one area"),

    ("Adjustment", "18-22", "Month 5", "TIE", "Company Values",
     "Demonstrate a company value in a visible way — make a decision or take an action explicitly aligned with values.",
     "Newcomer", "Situational", "Values understanding", "Value-aligned action taken and noted"),

    ("Adjustment", "18-22", "Month 5", "FIT", "RACI Matrix",
     "Identify one process where roles are unclear across teams. Propose a RACI clarification to your manager.",
     "Newcomer + Manager", "30 min", "RACI v2", "Process improvement proposed"),

    ("Adjustment", "18-22", "Month 5", "TIE", "Pulse Surveys",
     "Complete Month 5 check-in.",
     "Newcomer alone", "15 min", "Month 4 check-in", "Month 5 check-in submitted"),

    ("Adjustment", "23-26", "Month 6", "ACE", "Performance Appraisal",
     "Mid-year performance review. Present your achievements, development progress, and goals for H2.",
     "Newcomer + Manager + HR", "1 hour", "Ongoing performance data", "Mid-year review completed"),

    ("Adjustment", "23-26", "Month 6", "ACE", "Skills Matrix",
     "Update skills self-assessment. Compare to Day 5 baseline. Celebrate progress. Identify H2 development priorities.",
     "Newcomer + Manager", "30 min", "Dev plan progress", "Updated skills matrix; H2 priorities set"),

    ("Adjustment", "23-26", "Month 6", "TIE", "Employee Communities",
     "Take an active role in one community: present, organise an event, or mentor a newer newcomer.",
     "Newcomer + Community", "1-2 hours", "Community membership", "Active contributor to one community"),

    ("Adjustment", "23-26", "Month 6", "FIT", "Org Chart",
     "Can you navigate any decision — budget, hiring, escalation — to the right person without asking? Self-test.",
     "Newcomer alone", "15 min", "Stakeholder map", "Full org navigation confidence"),

    ("Adjustment", "23-26", "Month 6", "TIE", "Pulse Surveys",
     "Complete Month 6 check-in. Adjustment phase milestone — halfway through the journey.",
     "Newcomer alone", "15 min", "Month 5 check-in", "Month 6 check-in submitted; Adjustment closed"),

    # ═══════════════════════════════════════════════════════════
    # STABILIZATION — Month 7-9 (Days 181-270)
    # ═══════════════════════════════════════════════════════════
    ("Stabilization", "27-30", "Month 7", "FIT", "Strategic Plan",
     "Lead a cross-functional initiative aligned with company strategy. Own the outcome, not just the tasks.",
     "Newcomer + Cross-functional", "Ongoing", "Strategic initiative proposed", "Cross-functional initiative led"),

    ("Stabilization", "27-30", "Month 7", "ACE", "Training Plan",
     "Mentor a newer team member on a tool or process. Teaching consolidates your own mastery.",
     "Newcomer + Junior colleague", "1 hour/week", "Process expertise", "Mentoring relationship started"),

    ("Stabilization", "27-30", "Month 7", "TIE", "Buddy / Mentor",
     "Buddy relationship transitions to peer friendship. No more formal check-ins unless you want them.",
     "Newcomer + Buddy", "Casual", "6 months of buddy relationship", "Buddy is now a trusted colleague"),

    ("Stabilization", "27-30", "Month 7", "TIE", "Pulse Surveys",
     "Complete Month 7 check-in.",
     "Newcomer alone", "15 min", "Month 6 check-in", "Month 7 check-in submitted"),

    ("Stabilization", "31-35", "Month 8", "ACE", "SOPs & Playbooks",
     "Propose and implement a process improvement that affects the team. Get buy-in, execute, measure impact.",
     "Newcomer + Team", "2-4 hours", "Process expertise", "Process improvement implemented"),

    ("Stabilization", "31-35", "Month 8", "FIT", "KPIs / OKRs",
     "Q3 KPI review. Are you consistently hitting targets? Discuss promotion or role expansion with manager.",
     "Newcomer + Manager", "30 min", "Mid-year review", "Q3 review; career development discussed"),

    ("Stabilization", "31-35", "Month 8", "TIE", "Team Rituals",
     "You are now a cultural carrier. Help onboard any new person joining the team — show them the rituals.",
     "Newcomer + New joiner", "30 min", "Ritual participation + leadership", "Paid it forward to a newer newcomer"),

    ("Stabilization", "31-35", "Month 8", "TIE", "Pulse Surveys",
     "Complete Month 8 check-in.",
     "Newcomer alone", "15 min", "Month 7 check-in", "Month 8 check-in submitted"),

    ("Stabilization", "36-39", "Month 9", "ACE", "Performance Appraisal",
     "Prepare for year-end review: compile achievements, evidence, 360 feedback if available.",
     "Newcomer alone", "1 hour", "All performance data", "Year-end review preparation complete"),

    ("Stabilization", "36-39", "Month 9", "FIT", "Job Description",
     "Revisit your job description. Has your role grown beyond it? Discuss scope update or re-leveling with manager.",
     "Newcomer + Manager", "30 min", "90-day role review", "Role scope confirmed or updated"),

    ("Stabilization", "36-39", "Month 9", "TIE", "Company Values",
     "Can you explain the company values to a stranger convincingly? Could you give 3 real examples from your experience?",
     "Newcomer self-reflection", "15 min", "All values activities", "Deep values internalisation confirmed"),

    ("Stabilization", "36-39", "Month 9", "TIE", "Pulse Surveys",
     "Complete Month 9 check-in. Stabilization milestone — you should feel solid.",
     "Newcomer alone", "15 min", "Month 8 check-in", "Month 9 check-in submitted; Stabilization closed"),

    # ═══════════════════════════════════════════════════════════
    # EMBEDDING — Month 10-12 (Days 271-365)
    # ═══════════════════════════════════════════════════════════
    ("Embedding", "40-44", "Month 10", "FIT", "Strategic Plan",
     "Contribute to next year's planning. Propose goals, initiatives, or resource needs based on your experience.",
     "Newcomer + Manager + Team", "2 hours", "Cross-functional initiative", "Strategic contribution to next year's plan"),

    ("Embedding", "40-44", "Month 10", "ACE", "Skills Matrix",
     "Final skills assessment: compare to Day 5 baseline. You should have no critical gaps remaining.",
     "Newcomer + Manager", "30 min", "H2 development", "Skills matrix shows no critical gaps"),

    ("Embedding", "40-44", "Month 10", "TIE", "Employee Communities",
     "You're now a connector. Introduce two people from different departments who should know each other.",
     "Newcomer + Network", "15 min", "Broad network", "Network bridging — connected others"),

    ("Embedding", "40-44", "Month 10", "TIE", "Pulse Surveys",
     "Complete Month 10 check-in.",
     "Newcomer alone", "15 min", "Month 9 check-in", "Month 10 check-in submitted"),

    ("Embedding", "45-48", "Month 11", "ACE", "Performance Appraisal",
     "Year-end performance review. Present full year narrative: from newcomer to contributor. Receive formal rating.",
     "Newcomer + Manager + HR", "1 hour", "Review preparation", "Annual performance review completed"),

    ("Embedding", "45-48", "Month 11", "FIT", "KPIs / OKRs",
     "Set next year's OKRs with manager. You're now planning as an insider, not a newcomer.",
     "Newcomer + Manager", "45 min", "All KPI reviews", "Next year's OKRs set"),

    ("Embedding", "45-48", "Month 11", "TIE", "Buddy / Mentor",
     "Volunteer to be a buddy for the next newcomer joining your team.",
     "Newcomer + HR", "15 min", "Full buddy experience", "Signed up as future buddy"),

    ("Embedding", "45-48", "Month 11", "TIE", "Pulse Surveys",
     "Complete Month 11 check-in.",
     "Newcomer alone", "15 min", "Month 10 check-in", "Month 11 check-in submitted"),

    ("Embedding", "49-52", "Month 12", "FIT", "Job Description",
     "Final role reflection: Write a paragraph on how your understanding of the role evolved over 12 months.",
     "Newcomer alone", "20 min", "All FIT activities", "12-month role evolution narrative"),

    ("Embedding", "49-52", "Month 12", "ACE", "Training Plan",
     "Write a brief 'What I wish I'd known' guide for future newcomers in your role. Share with HR.",
     "Newcomer alone", "30 min", "Full 12-month experience", "Newcomer guide created for successors"),

    ("Embedding", "49-52", "Month 12", "TIE", "Team Rituals",
     "You no longer think about rituals — you live them. Reflect: Which ritual matters most to me? Why?",
     "Newcomer self-reflection", "10 min", "All ritual participation", "Rituals internalised"),

    ("Embedding", "49-52", "Month 12", "TIE", "Pulse Surveys",
     "Complete final Month 12 check-in. This closes your socialization journey. Celebrate.",
     "Newcomer alone", "15 min", "Month 11 check-in", "Journey complete — all 12 check-ins submitted"),

    ("Embedding", "52", "Day 365", "FIT", "Strategic Plan",
     "You are no longer new. You are part of the fabric. Final meeting with manager: reflect on the journey, look ahead.",
     "Newcomer + Manager", "30 min", "Everything", "Socialization journey officially closed"),
]

# ── Write data ──
for row_idx, act in enumerate(activities, 2):
    for col_idx, val in enumerate(act, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = WRAP
        cell.border = THIN
        cell.font = Font(name="Segoe UI", size=9)

    # Color the dimension column
    dim = act[3]
    if dim in DIM_FILLS:
        ws.cell(row=row_idx, column=4).fill = DIM_FILLS[dim]

    # Color the phase column
    phase = act[0]
    if phase in PHASE_FILLS:
        ws.cell(row=row_idx, column=1).fill = PHASE_FILLS[phase]

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:J{len(activities) + 1}"

# ── Summary sheet ──
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 12

# Count activities per phase per dimension
from collections import Counter
phase_dim = Counter()
for a in activities:
    phase_dim[(a[0], a[3])] += 1

phases_order = ["Pre-arrival", "Arrival", "Integration", "Adjustment", "Stabilization", "Embedding"]
dims = ["FIT", "ACE", "TIE"]

ws2.cell(row=1, column=1, value="Phase").font = Font(bold=True)
for ci, d in enumerate(dims, 2):
    ws2.cell(row=1, column=ci, value=d).font = Font(bold=True)
ws2.cell(row=1, column=5, value="Total").font = Font(bold=True)

for ri, phase in enumerate(phases_order, 2):
    ws2.cell(row=ri, column=1, value=phase)
    total = 0
    for ci, d in enumerate(dims, 2):
        count = phase_dim.get((phase, d), 0)
        ws2.cell(row=ri, column=ci, value=count)
        total += count
    ws2.cell(row=ri, column=5, value=total)

ws2.cell(row=len(phases_order) + 2, column=1, value="TOTAL").font = Font(bold=True)
for ci, d in enumerate(dims, 2):
    total = sum(phase_dim.get((p, d), 0) for p in phases_order)
    ws2.cell(row=len(phases_order) + 2, column=ci, value=total).font = Font(bold=True)
ws2.cell(row=len(phases_order) + 2, column=5, value=len(activities)).font = Font(bold=True)

out = r"C:\Users\bolve\projects\socialization-app\Socialization_Activity_Calendar.xlsx"
wb.save(out)
print(f"Saved {len(activities)} activities to {out}")
