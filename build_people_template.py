"""
Build a sample People import Excel template.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "People"

HEADER_FONT = Font(name="Segoe UI", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
NEWCOMER_FILL = PatternFill(start_color="EEEEF5", end_color="EEEEF5", fill_type="solid")
THIN = Border(
    left=Side(style="thin", color="E2E0DA"), right=Side(style="thin", color="E2E0DA"),
    top=Side(style="thin", color="E2E0DA"), bottom=Side(style="thin", color="E2E0DA"),
)

headers = [
    ("Name", 25), ("Email", 30), ("Role/Position", 25), ("Department", 18),
    ("Reports To", 30), ("Is Newcomer", 12), ("Start Date", 14), ("Buddy Email", 30), ("Password", 15),
]
for col, (h, w) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN
    ws.column_dimensions[chr(64 + col)].width = w

# Sample data - mix of existing employees and newcomers
people = [
    ("Elena Rodriguez", "elena@meridian.demo", "CEO", "Executive", "", "No", "", "", "demo123"),
    ("Claire Bennett", "claire@meridian.demo", "VP Marketing", "Marketing", "elena@meridian.demo", "No", "", "", "demo123"),
    ("Ravi Sharma", "ravi@meridian.demo", "VP Engineering", "Engineering", "elena@meridian.demo", "No", "", "", "demo123"),
    ("James Okafor", "james@meridian.demo", "Marketing Strategist", "Marketing", "claire@meridian.demo", "No", "", "", "demo123"),
    ("Nina Johansson", "hr@meridian.demo", "HR Director", "Human Resources", "elena@meridian.demo", "No", "", "", "demo123"),
    ("Sofia Martinez", "sofia@meridian.demo", "Sr. Marketing Manager", "Marketing", "claire@meridian.demo", "Yes", "2026-03-03", "james@meridian.demo", "demo123"),
    ("Daniel Cruz", "daniel@meridian.demo", "Data Analyst", "Analytics", "ravi@meridian.demo", "Yes", "2026-02-01", "", "demo123"),
    ("Yuki Tanaka", "yuki@meridian.demo", "Product Designer", "Product", "claire@meridian.demo", "Yes", "2026-01-15", "", "demo123"),
]

for i, row in enumerate(people, start=2):
    for col, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.border = THIN
        cell.alignment = Alignment(vertical="top")
    # Highlight newcomer rows
    if row[5] == "Yes":
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).fill = NEWCOMER_FILL

output = "People_Import_Template.xlsx"
wb.save(output)
print(f"Done! Template saved to {output}")
print(f"  {len(people)} sample records ({sum(1 for p in people if p[5] == 'Yes')} newcomers)")
